import openpyxl

path = 'E:\FORUM\Project\CE2020\SatelliteData.xlsx'

def satSearch(catalogNumber, path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    result = {}

    for row_number in range(1, sheet.max_row + 1):
        catalog = sheet.cell(row = row_number, column = 3).value
        if catalog == catalogNumber:
            result.update({'Name' : sheet.cell(row = row_number, column = 2).value})
            result.update({'Designator' : sheet.cell(row = row_number, column = 5).value})
            result.update({'Epoch' : sheet.cell(row = row_number, column = 6).value})
            result.update({'Inclination' : sheet.cell(row = row_number, column = 7).value})
            result.update({'RAAN' : sheet.cell(row = row_number, column = 8).value})
            e = sheet.cell(row = row_number, column = 9).value
            e = str(e)
            e = '0.' + e
            result.update({'Eccentricity' : e})
            result.update({'Perigee' : sheet.cell(row = row_number, column = 10).value})
            result.update({'Anomaly' : sheet.cell(row = row_number, column = 11).value})
            result.update({'Motion' : sheet.cell(row = row_number, column = 12).value})

            break
        else:
            result = {}
        
    return result 

def commitData(Epoch,ID,Date,Time,RA,Dec):
    wb = openpyxl.load_workbook('E:\FORUM\Project\CE2020\Analyse.xlsx')
    sheet = wb.active
    last_data_row = len(list(sheet.rows))
    last_row = last_data_row + 1
    sheet.cell(row=last_row,column = 1).value = Epoch
    sheet.cell(row=last_row,column = 2).value = ID

    sheet.cell(row=last_row,column = 4).value = Date
    sheet.cell(row=last_row,column = 5).value = Time

    sheet.cell(row=last_row,column = 7).value = RA
    sheet.cell(row=last_row,column = 8).value = Dec
    wb.save('E:\FORUM\Project\CE2020\Analyse.xlsx')
    wb.close()
    print('END DATA COMMIT')




