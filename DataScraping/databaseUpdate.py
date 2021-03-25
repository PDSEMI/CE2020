import openpyxl


######### EXCEL #########
#workbook = xlsxwriter.Workbook("SatelliteData.xls")
workbook = openpyxl.load_workbook("E:\FORUM\Project\CE2020\SatelliteData.xlsx")
worksheet = workbook.active

Header = ['ID', 'NAME', 'CATALOG', 'ELSET CLASSIFICATION', 'INTERNATIONAL DESIGNATOR', 'EPOCH', 'INCLINATION', 'RA OF ASC_NODE','ECCENTRICITY', 'ARG OF PERIGEE', 'MEAN ANOMALY', 'MOTION']
i = 1
for column_name in Header:
    worksheet.cell(row = 1, column = i).value = column_name
    i = i+1


######### END OF DATABASE INITIALIZE #########
class Satellite:
    def __init__(self):
        self.Name = None
        self.CatNum = None
        self.Class = None
        self.Desig = None
        self.Epoch = None
        self.Incl = None
        self.RA = None
        self.E = None
        self.Perigee = None
        self.Anomaly = None
        self.Motion = None

def to_sat_data(sat):
    sat_data =  [sat.Name,sat.CatNum,sat.Class,sat.Desig,sat.Epoch,sat.Incl,sat.RA,sat.E,sat.Perigee,sat.Anomaly,sat.Motion]
    return sat_data
   
######### INSERT DATA ##########

data_file = r"E:\FORUM\Project\CE2020\DataScraping\3LE.txt"
data = open(data_file,"r")
i = 0
sat = Satellite()
row = 2
for line in data:
    column = 1
    i = i+1
    if i%3 == 1:
        name = line[0:24].strip()
        sat.Name = name
    elif i%3 == 2:
        sat.CatNum = line[2:7].strip()
        sat.Class = line[7].strip()
        sat.Desig = line[9:17].strip()
        sat.Epoch = line[17:32].strip()
    elif i%3 == 0:
        sat.Incl = line[8:17].strip()
        sat.RA = line[17:25].strip()
        sat.E = line[26:33].strip()
        sat.Perigee = line[34:42].strip()
        sat.Anomaly = line[43:51].strip()
        sat.Motion = line[52:63].strip()
        sat_data = to_sat_data(sat)
        worksheet.cell(row = row, column = 1).value = row - 1
        for data in sat_data:
            worksheet.cell(row = row, column = column+1).value = data
            column = column + 1 
        row = row +1

workbook.save("E:\FORUM\Project\CE2020\SatelliteData.xlsx")
print("END SECTION")

